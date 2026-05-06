"""
Node: read_input — reads the Excel file and populates initial state.
"""

from __future__ import annotations
from openpyxl import load_workbook
import config
from graph.state import BookState


def read_input(state: BookState) -> BookState:
    """
    Read the Excel workbook at INPUT_EXCEL_PATH.
    Expected columns (row 1 = headers, row 2 = values):
        title, notes_on_outline_before, status_outline_notes,
        notes_on_outline_after, chapter_notes_status, chapter_notes,
        final_review_notes_status, final_review_notes
    """
    path = config.INPUT_EXCEL_PATH
    if not path.exists():
        return {
            **state,
            "status": "error",
            "error": f"Input file not found: {path}",
            "message": f"❌ Input file not found at {path}",
        }

    wb = load_workbook(str(path), data_only=True)
    ws = wb.active

    # Build header → column-index map from row 1
    headers: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = col_idx

    def _cell(name: str) -> str:
        """Get a cell value from row 2 by header name."""
        col = headers.get(name)
        if col is None:
            return ""
        val = ws.cell(row=2, column=col).value
        return str(val).strip() if val is not None else ""

    title = _cell("title")
    if not title:
        return {
            **state,
            "status": "error",
            "error": "Title is missing in the input Excel file.",
            "message": "❌ Title is mandatory but missing in input file.",
        }

    return {
        **state,
        "title": title,
        "notes_on_outline_before": _cell("notes_on_outline_before"),
        "notes_on_outline_after": _cell("notes_on_outline_after"),
        "status_outline_notes": _cell("status_outline_notes").lower(),
        "chapter_notes_status": _cell("chapter_notes_status").lower(),
        "chapter_notes": _cell("chapter_notes"),
        "final_review_notes_status": _cell("final_review_notes_status").lower(),
        "final_review_notes": _cell("final_review_notes"),
        "status": "running",
        "message": f"✅ Input loaded — Title: '{title}'",
    }
