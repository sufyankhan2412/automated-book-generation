"""
Main entry point — runs the LangGraph book-generation workflow.

Usage:
    python main.py                  # Normal run (reads from Excel)
    python main.py --resume         # Resume from where it paused
    python main.py --setup-db       # Print SQL setup instructions
"""

from __future__ import annotations
import sys
from pathlib import Path

from graph.workflow import build_workflow
from graph.state import BookState


def _try_resume() -> BookState:
    """
    Check Supabase for an existing in-progress book.
    If found, restore state so the workflow can resume.
    """
    from db.supabase_client import get_client
    from db import models
    from openpyxl import load_workbook
    import config

    client = get_client()

    # Find a book that isn't completed
    result = client.table("books").select("*").neq("book_output_status", "completed").order("created_at", desc=True).limit(1).execute()
    if not result.data:
        return {}

    book = result.data[0]
    book_id = book["id"]
    title = book["title"]

    # Get the latest outline
    outline_row = models.get_latest_outline(book_id)
    if not outline_row:
        return {}

    # Get existing chapters
    chapters = models.get_all_chapters(book_id)
    last_chapter = max((c["chapter_number"] for c in chapters), default=0)

    # Re-read Excel for latest user inputs (notes, statuses)
    path = config.INPUT_EXCEL_PATH
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = col_idx

    def _cell(name):
        col = headers.get(name)
        if col is None:
            return ""
        val = ws.cell(row=2, column=col).value
        return str(val).strip() if val is not None else ""

    # Parse chapter titles from outline
    from graph.nodes.outline_generator import _parse_chapter_titles
    chapter_titles = _parse_chapter_titles(outline_row["outline_text"])

    # Resume from after the last completed chapter
    resume_chapter = last_chapter  # continue from next chapter
    print(f"▶️  Resuming: {last_chapter}/{len(chapter_titles)} chapters done, continuing to next")

    state: BookState = {
        "title": title,
        "book_id": book_id,
        "outline": outline_row["outline_text"],
        "outline_version": outline_row["version"],
        "outline_chapters": chapter_titles,
        "total_chapters": len(chapter_titles),
        "current_chapter": resume_chapter,
        "chapters_completed": False,
        "notes_on_outline_before": _cell("notes_on_outline_before"),
        "notes_on_outline_after": _cell("notes_on_outline_after"),
        "status_outline_notes": "no_notes_needed",  # outline already done
        "chapter_notes_status": "no_notes_needed",
        "chapter_notes": "",
        "final_review_notes_status": _cell("final_review_notes_status").lower(),
        "final_review_notes": _cell("final_review_notes"),
        "status": "running",
        "message": f"▶️  Resumed book '{title}' — {last_chapter}/{len(chapter_titles)} chapters done",
    }
    return state


def main():
    # ── Handle CLI flags ─────────────────────────────
    if "--setup-db" in sys.argv:
        sql_path = Path(__file__).parent / "db" / "setup.sql"
        print("Run this SQL in your Supabase SQL Editor:\n")
        print(sql_path.read_text(encoding="utf-8"))
        return

    # ── Build & run workflow ─────────────────────────
    print("=" * 60)
    print("  📖 Automated Book Generation System")
    print("=" * 60)

    workflow = build_workflow()

    # Check if we should resume an existing book
    initial_state: BookState = {}
    if "--resume" in sys.argv:
        print("\n🔍 Checking for in-progress book to resume...")
        initial_state = _try_resume()
        if initial_state:
            print(f"  📚 Found: '{initial_state.get('title', '')}'")
            print(f"  Progress: {initial_state.get('current_chapter', 0)}/{initial_state.get('total_chapters', 0)} chapters")
        else:
            print("  No in-progress book found. Starting fresh.\n")

    print("\n🚀 Starting workflow...\n")

    # Run the graph
    final_state = workflow.invoke(
        initial_state,
        config={"recursion_limit": 200},
    )

    # ── Print final status ───────────────────────────
    print("\n" + "=" * 60)
    print("  WORKFLOW RESULT")
    print("=" * 60)
    print(f"  Status:  {final_state.get('status', 'unknown')}")
    print(f"  Message: {final_state.get('message', 'N/A')}")

    if final_state.get("error"):
        print(f"  Error:   {final_state['error']}")

    if final_state.get("book_id"):
        print(f"  Book ID: {final_state['book_id']}")

    if final_state.get("status") == "paused":
        print(f"\n  💡 To resume from where you left off:")
        print(f"     Run: python main.py --resume")

    print("=" * 60)

    return final_state


if __name__ == "__main__":
    main()
